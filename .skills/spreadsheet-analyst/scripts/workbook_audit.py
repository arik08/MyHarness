#!/usr/bin/env python3
"""Summarize spreadsheet structure and obvious quality issues."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from statistics import mean
from typing import Any

ERROR_LITERALS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}


def _json(data: dict[str, Any]) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2))


def _read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(errors="replace")


def audit_csv(path: Path, delimiter: str | None) -> dict[str, Any]:
    text = _read_text(path)
    sample = text[:4096]
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(sample).delimiter
        except csv.Error:
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","

    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    widths = [len(row) for row in rows]
    nonempty_rows = [row for row in rows if any(cell.strip() for cell in row)]
    width_counts: dict[int, int] = {}
    for width in widths:
        width_counts[width] = width_counts.get(width, 0) + 1

    return {
        "path": str(path),
        "kind": path.suffix.lower().lstrip("."),
        "delimiter": delimiter,
        "rows": len(rows),
        "nonempty_rows": len(nonempty_rows),
        "columns_min": min(widths) if widths else 0,
        "columns_max": max(widths) if widths else 0,
        "columns_avg": round(mean(widths), 2) if widths else 0,
        "ragged_rows": sum(1 for width in widths if widths and width != max(widths)),
        "width_counts": width_counts,
        "preview": rows[:5],
    }


def audit_workbook(path: Path, max_cells: int) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return {
            "path": str(path),
            "kind": path.suffix.lower().lstrip("."),
            "error": "openpyxl is required to audit Excel workbooks",
            "details": str(exc),
        }

    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, data_only=False, keep_vba=keep_vba)
    sheets: list[dict[str, Any]] = []
    totals = {
        "formulas": 0,
        "error_literals": 0,
        "comments": 0,
        "tables": 0,
        "merged_ranges": 0,
        "scanned_cells": 0,
    }

    for sheet in workbook.worksheets:
        formulas = 0
        errors: list[str] = []
        comments = 0
        scanned = 0
        truncated = False

        for row in sheet.iter_rows():
            for cell in row:
                scanned += 1
                if scanned > max_cells:
                    truncated = True
                    break
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                if value in ERROR_LITERALS or cell.data_type == "e":
                    errors.append(cell.coordinate)
                if cell.comment is not None:
                    comments += 1
            if truncated:
                break

        table_count = len(getattr(sheet, "tables", {}))
        merged_count = len(sheet.merged_cells.ranges)
        totals["formulas"] += formulas
        totals["error_literals"] += len(errors)
        totals["comments"] += comments
        totals["tables"] += table_count
        totals["merged_ranges"] += merged_count
        totals["scanned_cells"] += min(scanned, max_cells)

        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "formulas": formulas,
                "error_literals": len(errors),
                "error_locations": errors[:25],
                "comments": comments,
                "tables": table_count,
                "merged_ranges": merged_count,
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "auto_filter": str(sheet.auto_filter.ref) if sheet.auto_filter and sheet.auto_filter.ref else None,
                "scan_truncated": truncated,
            }
        )

    return {
        "path": str(path),
        "kind": path.suffix.lower().lstrip("."),
        "sheet_count": len(workbook.worksheets),
        "sheets": sheets,
        "totals": totals,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit spreadsheet structure and obvious errors.")
    parser.add_argument("path", help="Path to .xlsx, .xlsm, .csv, or .tsv file")
    parser.add_argument("--delimiter", help="Delimiter override for CSV/TSV files")
    parser.add_argument("--max-cells", type=int, default=200000, help="Maximum cells to scan per sheet")
    args = parser.parse_args()

    path = Path(args.path).expanduser().resolve()
    if not path.exists():
        _json({"path": str(path), "error": "file not found"})
        return 2

    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        _json(audit_csv(path, args.delimiter))
        return 0
    if suffix in {".xlsx", ".xlsm"}:
        _json(audit_workbook(path, args.max_cells))
        return 0

    _json({"path": str(path), "error": f"unsupported spreadsheet extension: {suffix}"})
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
